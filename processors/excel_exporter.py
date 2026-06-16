from __future__ import annotations

import io
from typing import List, Dict, Any

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def _format_sheet(ws):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_len = 10
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, min(len(value), 60))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2


def build_excel(
    movements: List[Dict[str, Any]],
    ocr_pages: List[Dict[str, Any]],
    source_filename: str = "",
) -> bytes:
    output = io.BytesIO()

    df_mov = pd.DataFrame(movements)
    if df_mov.empty:
        df_mov = pd.DataFrame(columns=[
            "Dia", "Descripcion", "Referencia/Serial", "Cargo",
            "Abono", "Saldo", "Pagina", "Observacion", "Linea OCR"
        ])

    df_ocr = pd.DataFrame([
        {"Pagina": p["page"], "Texto OCR": p["text"]}
        for p in ocr_pages
    ])

    df_info = pd.DataFrame([
        {"Campo": "Archivo origen", "Valor": source_filename},
        {"Campo": "Movimientos detectados", "Valor": len(df_mov)},
        {"Campo": "Motor OCR", "Valor": "Tesseract local en Streamlit Cloud"},
        {"Campo": "Nota", "Valor": "Revisar filas con observación. Fase 1 sin costo."},
    ])

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_info.to_excel(writer, sheet_name="INFO", index=False)
        df_mov.to_excel(writer, sheet_name="MOVIMIENTOS", index=False)
        df_ocr.to_excel(writer, sheet_name="OCR_TEXTO", index=False)

        wb = writer.book

        for sheet_name in ["INFO", "MOVIMIENTOS", "OCR_TEXTO"]:
            ws = wb[sheet_name]
            _format_sheet(ws)

        ws_mov = wb["MOVIMIENTOS"]
        money_cols = ["Cargo", "Abono", "Saldo"]
        headers = [cell.value for cell in ws_mov[1]]

        for col_name in money_cols:
            if col_name in headers:
                col_idx = headers.index(col_name) + 1
                for row in range(2, ws_mov.max_row + 1):
                    ws_mov.cell(row=row, column=col_idx).number_format = '$#,##0.00'

        if "Observacion" in headers:
            obs_col = headers.index("Observacion") + 1
            fill = PatternFill("solid", fgColor="FFF2CC")
            for row in range(2, ws_mov.max_row + 1):
                if ws_mov.cell(row=row, column=obs_col).value:
                    for col in range(1, ws_mov.max_column + 1):
                        ws_mov.cell(row=row, column=col).fill = fill

    output.seek(0)
    return output.getvalue()
